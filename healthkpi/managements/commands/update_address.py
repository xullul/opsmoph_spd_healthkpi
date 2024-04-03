import io
import requests
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from healthkpi.models import HealthZone, Region, Province, District, SubDistrict


class Command(BaseCommand):
    excel_url = r'https://stat.bora.dopa.go.th/dload/ccaatt.xlsx'
    more_information_url = r'https://stat.bora.dopa.go.th/stat/statnew/statMenu/newStat/ccaa.php'
    help = (
        f'load excel data from "{excel_url}" into io stream '
        'and perform update Province, District and SubDistrict tables, '
        f'for more information, checkout "{more_information_url}"'
    )
    
    def handle(self, *args, **options) -> None:
        response = requests.get(self.excel_url)
        
        if not response.ok:
            raise CommandError(
                f'Failed to request excel file from "{self.excel_url}":\n'
                f'status code: {response.status_code}\n'
                f'{response.text}'
            )
            
        excel = io.BytesIO(response.content)
        wb: Workbook = load_workbook(excel)
        ws: Worksheet = wb.active
        start_row: int = 1
        stop_row: int = 1
        
        # determine which row to start
        for row_number in range(1, ws.max_row + 1):
            row = ws.cell(row=row_number, column=1).value
            if isinstance(row, str) and (len(row) == 8 or len(row) == 6) and row.isnumeric():
                start_row = row_number
                break
        
        # determine which row to stop
        for row_number in range(ws.max_row, 0, -1):
            row = ws.cell(row=row_number, column=1).value
            if isinstance(row, str) and (len(row) == 8 or len(row) == 6) and row.isnumeric():
                stop_row = row_number
                break
        
        total_rows: int = stop_row - start_row + 1
        
        # column 1 is address code
        # column 2 is address name
        # column 3 is address english name
        # column 4 is deploy date (not used in this case)
        
        # find and update province
        for count, row_number in enumerate(range(start_row, stop_row + 1), start=1):
            code: str = ws.cell(row=row_number, column=1).value[0:6]
            province_code = code[0:2]
            district_code = code[2:4]
            subdistrict_code = code[4:6]
            name = ws.cell(row=row_number, column=2).value
            en_name = ws.cell(row=row_number, column=3)
            if province_code != '00' and district_code == '00' and subdistrict_code == '00':
                Province.objects.update_or_create()